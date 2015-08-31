import os

__author__ = 'Killx'

import tkinter
import csv
from tkinter import *
from tkinter import ttk, tix
from tkinter.tix import ScrolledWindow
from openpyxl import Workbook
from openpyxl import *
from openpyxl.styles import Border, Side, Alignment, Protection, Font, colors, Style
from openpyxl.styles.borders import Border, Side
from openpyxl.cell import *
import tkinter.messagebox as tkMessageBox

global i, b, c, sem_list, d, subcon, semcount, semname, semyear, subject, subjinc, subjcode, subjname, cha, chp, gr, gp, w, first, last, sd, ssn, w, deg, major, matric, sex, master
semyear = [''] * 14
semname = [''] * 14
subcon = [''] * 14
subject = [''] * 100
subjcode = [''] * 100
subjname = [''] * 100
cha = [''] * 100
chp = [''] * 100
gr = [''] * 100
gp = [''] * 100
semcount = -1
subjinc = -1


# as it is
class AutocompleteEntry(tkinter.Entry):
    """
       D Subclass of tkinter.Entry that features autocompletion.
      To enable autocompletion use set_completion_list(list) to define   H/
       R a list of possible strings to hit.
        To cycle through hits use down and up arrow keys. U
        V  Autocomplete help
        """

    def set_completion_list(self, completion_list):
        self._completion_list = sorted(completion_list, key=str.lower)  # Work with a sorted list
        self._hits = []
        self._hit_index = 0
        self.position = 0
        self.bind('<KeyRelease>', self.handle_keyrelease)

    def autocomplete(self, delta=0):
        """autocomplete the Entry, delta may be 0/1/-1 to cycle through possible hits"""
        if delta:  # need to delete selection otherwise we would fix the current position
            self.delete(self.position, tkinter.END)
        else:  # set position to end so selection starts where textentry ended
            self.position = len(self.get())
        # collect hits
        _hits = []
        for element in self._completion_list:
            if element.lower().startswith(self.get().lower()):  # Match case-insensitively
                _hits.append(element)
        # if we have a new hit list, keep this in mind
        if _hits != self._hits:
            self._hit_index = 0
            self._hits = _hits
        # only allow cycling if we are in a known hit list
        if _hits == self._hits and self._hits:
            self._hit_index = (self._hit_index + delta) % len(self._hits)
        # now finally perform the auto completion
        if self._hits:
            self.delete(0, tkinter.END)
            self.insert(0, self._hits[self._hit_index])
            self.select_range(self.position, tkinter.END)

    def handle_keyrelease(self, event):
        """event handler for the keyrelease event on this widget"""
        if event.keysym == "BackSpace":
            self.delete(self.index(tkinter.INSERT), tkinter.END)
            self.position = self.index(tkinter.END)
        if event.keysym == "Left":
            if self.position < self.index(tkinter.END):  # delete the selection
                self.delete(self.position, tkinter.END)
            else:
                self.position = self.position - 1  # delete one character
                self.delete(self.position, tkinter.END)
        if event.keysym == "Right":
            self.position = self.index(tkinter.END)  # go to end (no selection)
        if event.keysym == "Down":
            self.autocomplete(1)  # cycle to next hit
        if event.keysym == "Up":
            self.autocomplete(-1)  # cycle to previous hit
        if len(event.keysym) == 1:
            self.autocomplete()


# as it is
class AutocompleteCombobox(ttk.Combobox):
    def set_completion_list(self, completion_list):

        self._completion_list = sorted(completion_list, key=str.lower)
        self._hits = []
        self._hit_index = 0
        self.position = 0
        self.bind('<KeyRelease>', self.handle_keyrelease)
        self['values'] = self._completion_list

    def autocomplete(self, delta=0):

        if delta:
            self.delete(self.position, tkinter.END)
        else:
            self.position = len(self.get())

        _hits = []
        for element in self._completion_list:
            if element.lower().startswith(self.get().lower()):
                _hits.append(element)

        if _hits != self._hits:
            self._hit_index = 0
            self._hits = _hits

        if _hits == self._hits and self._hits:
            self._hit_index = (self._hit_index + delta) % len(self._hits)

        if self._hits:
            self.delete(0, tkinter.END)
            self.insert(0, self._hits[self._hit_index])
            self.select_range(self.position, tkinter.END)

    def handle_keyrelease(self, event):

        if event.keysym == "BackSpace":
            self.delete(self.index(tkinter.INSERT), tkinter.END)
            self.position = self.index(tkinter.END)
        if event.keysym == "Left":
            if self.position < self.index(tkinter.END):
                self.delete(self.position, tkinter.END)
            else:
                self.position = self.position - 1
                self.delete(self.position, tkinter.END)
        if event.keysym == "Right":
            self.position = self.index(tkinter.END)
        if len(event.keysym) == 1:
            self.autocomplete()


# as it is
def style_range(ws, cell_range, style=None):
    start_cell, end_cell = cell_range.split(':')
    start_coord = coordinate_from_string(start_cell)
    start_row = start_coord[1]
    start_col = column_index_from_string(start_coord[0])
    end_coord = coordinate_from_string(end_cell)
    end_row = end_coord[1]
    end_col = column_index_from_string(end_coord[0])

    for row in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            col = get_column_letter(col_idx)
            ws.cell('%s%s' % (col, row)).style = style


# ,modify if semcount>0 forget all  details for all subjects in range subjcount[semcount]-subjcount[semcount-1]
def callsem():
    global i, b, c, sem_list, d, subcon, semcount, semname, semyear, subject, subjinc, subjcode, subjname, cha, chp, gr, gp

    # destroy buttons to add new content
    b.destroy()
    c.destroy()
    d.destroy()

    semcount += 1
    subjinc += 1
    i += 1
    if (semcount > 1):

        i = 15
        semname[semcount - 1].grid_forget()
        semyear[semcount - 1].grid_forget()
        subcon[semcount - 1].grid_forget()
        a = int(subcon[semcount - 1].get())
        b = subjinc
        for x in range(a, b):
            subjcode[x].grid_forget()
            subjname[x].grid_forget()
            cha[x].grid_forget()
            chp[x].grid_forget()
            gr[x].grid_forget()
            gp[x].grid_forget()

        semname[semcount] = AutocompleteEntry(master)
        semname[semcount].set_completion_list(sem_list)
        semname[semcount].grid(row=13, column=2)
        semname[semcount].focus_set()
        semyear[semcount] = Entry(master)
        semyear[semcount].grid(row=13, column=3)
        subcon[semcount] = Entry(master)
        subcon[semcount].grid(row=13, column=5)
        subjcode[subjinc] = AutocompleteEntry(master)
        subjcode[subjinc].set_completion_list(code_list)
        subjcode[subjinc].grid(row=i, column=1, padx=2, pady=2)
        subjcode[subjinc].focus_set()

        subjname[subjinc] = AutocompleteEntry(master)
        subjname[subjinc].set_completion_list(subject_list)
        subjname[subjinc].grid(row=i, column=2, padx=2, pady=2)
        subjname[subjinc].focus_set()
        cha[subjinc] = Entry(master)
        cha[subjinc].grid(row=i, column=3, padx=2, pady=2)
        chp[subjinc] = Entry(master)
        chp[subjinc].grid(row=i, column=4, padx=2, pady=2)
        gr[subjinc] = Entry(master)
        gr[subjinc].grid(row=i, column=5, padx=2, pady=2)
        gp[subjinc] = Entry(master)
        gp[subjinc].grid(row=i, column=6, padx=2, pady=2)

    if (semcount == 1):
        i = 15
        semname[semcount - 1].grid_forget()
        semyear[semcount - 1].grid_forget()
        subcon[semcount - 1].grid_forget()
        for x in range(0, int(subcon[semcount - 1].get())):
            # z = int(x)
            subjcode[x].grid_forget()
            subjname[x].grid_forget()
            cha[x].grid_forget()
            chp[x].grid_forget()
            gr[x].grid_forget()
            gp[x].grid_forget()

        semname[semcount] = AutocompleteEntry(master)
        semname[semcount].set_completion_list(sem_list)
        semname[semcount].grid(row=13, column=2)
        semname[semcount].focus_set()
        semyear[semcount] = Entry(master)
        semyear[semcount].grid(row=13, column=3)
        subcon[semcount] = Entry(master)
        subcon[semcount].grid(row=13, column=5)
        subjcode[subjinc] = AutocompleteEntry(master)
        subjcode[subjinc].set_completion_list(code_list)
        subjcode[subjinc].grid(row=i, column=1, padx=2, pady=2)
        subjcode[subjinc].focus_set()

        subjname[subjinc] = AutocompleteEntry(master)
        subjname[subjinc].set_completion_list(subject_list)
        subjname[subjinc].grid(row=i, column=2, padx=2, pady=2)
        subjname[subjinc].focus_set()
        cha[subjinc] = Entry(master)
        cha[subjinc].grid(row=i, column=3, padx=2, pady=2)
        chp[subjinc] = Entry(master)
        chp[subjinc].grid(row=i, column=4, padx=2, pady=2)
        gr[subjinc] = Entry(master)
        gr[subjinc].grid(row=i, column=5, padx=2, pady=2)
        gp[subjinc] = Entry(master)
        gp[subjinc].grid(row=i, column=6, padx=2, pady=2)

    if (semcount == 0):
        Label(master, text="Semester").grid(row=13, column=1)
        semname[semcount] = AutocompleteEntry(master)
        semname[semcount].set_completion_list(sem_list)
        semname[semcount].grid(row=13, column=2)
        semname[semcount].focus_set()
        semyear[semcount] = Entry(master)
        semyear[semcount].grid(row=13, column=3)
        Label(master, text="Number of Subjects").grid(row=13, column=4)
        subcon[semcount] = Entry(master)
        subcon[semcount].grid(row=13, column=5)

        Label(master, text="Course id").grid(row=14, column=1)
        Label(master, text="Course name").grid(row=14, column=2)
        Label(master, text="Credit Hours Attempted").grid(row=14, column=3)
        Label(master, text="Credit Hours Passed").grid(row=14, column=4)
        Label(master, text="Grade").grid(row=14, column=5)
        Label(master, text="Quality Points").grid(row=14, column=6)
        i = 15

        subjcode[subjinc] = AutocompleteEntry(master)
        subjcode[subjinc].set_completion_list(code_list)
        subjcode[subjinc].grid(row=i, column=1, padx=2, pady=2)
        subjcode[subjinc].focus_set()

        subjname[subjinc] = AutocompleteEntry(master)
        subjname[subjinc].set_completion_list(subject_list)
        subjname[subjinc].grid(row=i, column=2, padx=2, pady=2)
        subjname[subjinc].focus_set()

        cha[subjinc] = Entry(master)
        cha[subjinc].grid(row=i, column=3, padx=2, pady=2)
        chp[subjinc] = Entry(master)
        chp[subjinc].grid(row=i, column=4, padx=2, pady=2)
        gr[subjinc] = Entry(master)
        gr[subjinc].grid(row=i, column=5, padx=2, pady=2)
        gp[subjinc] = Entry(master)
        gp[subjinc].grid(row=i, column=6, padx=2, pady=2)
    i += 1
    c = Button(master, text="Add Subject", command=callback)
    c.grid(row=i, column=1, padx=2, pady=2)
    b = Button(master, text="Add Semester", command=callsem)
    b.grid(row=i, column=2, padx=2, pady=2)
    d = Button(master, text="Enter Final Detail", command=calldegree)
    d.grid(row=i, column=4, padx=2, pady=2)


# to add subjects in current semester
def callback():
    global i, b, c, sem_list, d, subcon, semcount, semname, semyear, subject, subjinc, subjcode, subjname, cha, chp, gr, gp

    b.destroy()
    c.destroy()
    d.destroy()
    i += 1
    subjinc += 1
    subjcode[subjinc] = AutocompleteEntry(master)
    subjcode[subjinc].set_completion_list(code_list)
    subjcode[subjinc].grid(row=i, column=1, padx=2, pady=2)
    subjcode[subjinc].focus_set()

    subjname[subjinc] = AutocompleteEntry(master)
    subjname[subjinc].set_completion_list(subject_list)
    subjname[subjinc].grid(row=i, column=2, padx=2, pady=2)
    subjname[subjinc].focus_set()

    cha[subjinc] = Entry(master)
    cha[subjinc].grid(row=i, column=3, padx=2, pady=2)
    chp[subjinc] = Entry(master)
    chp[subjinc].grid(row=i, column=4, padx=2, pady=2)
    gr[subjinc] = Entry(master)
    gr[subjinc].grid(row=i, column=5, padx=2, pady=2)
    gp[subjinc] = Entry(master)
    gp[subjinc].grid(row=i, column=6, padx=2, pady=2)
    i += 1
    c = Button(master, text="Add Subject", command=callback)
    c.grid(row=i, column=1, padx=2, pady=2)
    b = Button(master, text="Add Semester", command=callsem)
    b.grid(row=i, column=2, padx=2, pady=2)
    d = Button(master, text="Enter Final Detail", command=calldegree)
    d.grid(row=i, column=4, padx=2, pady=2)


# to enter final details about the degree
def calldegree():
    global i, b, c, sem_list, d, subcon, semcount, semname, semyear, subject, subjinc, subjcode, subjname, cha, chp, gr, gp, da, datee, remark
    b.destroy()
    c.destroy()
    d.destroy()
    i += 1
    Label(master, text="Degree Awarded").grid(row=i, column=1)
    da = Entry(master)
    da.grid(row=i, column=2)
    Label(master, text="Date").grid(row=i, column=3)
    datee = Entry(master)
    datee.grid(row=i, column=4)
    i += 1
    Label(master, text="Remarks").grid(row=i, column=1)
    remark = Entry(master)
    remark.grid(row=i, column=2, sticky='w')
    i += 1
    gen = Button(master, text="Generate Transcript", command=generate)
    gen.grid(row=i, column=2)


def generate():
    global i, b, c, sem_list, d, subcon, semcount, remark, semname, semyear, subject, subjinc, subjcode, subjname, cha, chp, gr, gp, w, first, last, sd, ssn, w, deg, major, matric, sex, variable, acc, artnmus, bkg, cis, crj, eco, edu, eng, finnmgt, his, insnre, lan, law, mar, mgt, natsci, nur, psy, sec, spedra, socsci, tax, busele, laele, oc, hs, hsg, adg, sch, tc

    gender = sex.get()

    z = semcount

    t = 0

    wb = load_workbook(filename='dummy.xlsx')
    ws = wb.active
    name1 = first.get()
    lastnamee = last.get()
    nameforfile = name1 + "_" + lastnamee
    ws['B4'] = first.get() + " " + last.get()

    ws['B15'] = sd.get()
    ws['C3'] = ssn.get()
    campus = variable.get()
    ws['E3'] = deg.get()
    ws['S23'] = major.get()
    style_range(ws, 'QS23:AB23', Style(alignment=Alignment(horizontal='center'),
                                       border=Border(right=Side(border_style='thin', color=colors.BLACK))), )
    ws['S29'] = major.get()
    ws['G3'] = matric.get()
    style_range(ws, 'G3:I3', Style(alignment=Alignment(horizontal='center'),
                                   border=Border(left=Side(border_style='thin', color=colors.BLACK),
                                                 top=Side(border_style='thin', color=colors.BLACK),
                                                 bottom=Side(border_style='thin', color=colors.BLACK),
                                                 right=Side(border_style='thin', color=colors.BLACK), )), )

    ws['J3'] = sex.get()
    style_range(ws, 'J3:L3', Style(alignment=Alignment(horizontal='center'),
                                   border=Border(left=Side(border_style='thin', color=colors.BLACK),
                                                 top=Side(border_style='thin', color=colors.BLACK),
                                                 bottom=Side(border_style='thin', color=colors.BLACK),
                                                 right=Side(border_style='thin', color=colors.BLACK), )), )
    ws['E9'] = acc.get()
    ws['F9'] = artnmus.get()
    ws['G9'] = bkg.get()
    ws['H9'] = cis.get()
    ws['I9'] = crj.get()
    ws['J9'] = eco.get()
    ws['K9'] = edu.get()
    ws['L9'] = eng.get()
    ws['M9'] = finnmgt.get()
    ws['N9'] = his.get()
    ws['O9'] = insnre.get()
    ws['P9'] = lan.get()
    ws['Q9'] = law.get()
    ws['R9'] = mar.get()
    ws['S9'] = mgt.get()
    ws['T9'] = natsci.get()
    ws['U9'] = nur.get()
    ws['V9'] = psy.get()
    ws['W9'] = sec.get()
    ws['X9'] = spedra.get()
    ws['Y9'] = socsci.get()
    ws['Z9'] = tax.get()
    ws['AA9'] = busele.get()
    ws['AB9'] = laele.get()
    ws['B9'] = oc.get()
    style_range(ws, 'B9:C9', Style(alignment=Alignment(horizontal='center'),
                                   border=Border(left=Side(border_style='thin', color=colors.BLACK),
                                                 top=Side(border_style='thin', color=colors.BLACK),
                                                 bottom=Side(border_style='thin', color=colors.BLACK),
                                                 right=Side(border_style='thin', color=colors.BLACK), )), )
    ws['D9'] = adg.get()
    ws['D15'] = hsg.get()
    ws['C15'] = hs.get()
    ws['D3'] = sch.get()
    ws['Z10'] = tc.get()
    style_range(ws, 'Z10:AB10', Style(alignment=Alignment(horizontal='center'),
                                      border=Border(left=Side(border_style='thin', color=colors.BLACK),
                                                    top=Side(border_style='thin', color=colors.BLACK),
                                                    bottom=Side(border_style='thin', color=colors.BLACK),
                                                    right=Side(border_style='thin', color=colors.BLACK), )), )
    ws['Q34'] = remark.get()
    style_range(ws, 'Q34:AB34', Style(alignment=Alignment(horizontal='center'),
                                      border=Border(left=Side(border_style='thin', color=colors.BLACK),
                                                    right=Side(border_style='thin', color=colors.BLACK))), )
    style_range(ws, 'Q35:AB35', Style(alignment=Alignment(horizontal='center'),
                                      border=Border(left=Side(border_style='thin', color=colors.BLACK),
                                                    right=Side(border_style='thin', color=colors.BLACK))), )
    style_range(ws, 'Q36:AB36', Style(alignment=Alignment(horizontal='center'),
                                      border=Border(left=Side(border_style='thin', color=colors.BLACK),
                                                    right=Side(border_style='thin', color=colors.BLACK))), )
    style_range(ws, 'Q37:AB37', Style(alignment=Alignment(horizontal='center'),
                                      border=Border(left=Side(border_style='thin', color=colors.BLACK),
                                                    right=Side(border_style='thin', color=colors.BLACK))), )
    style_range(ws, 'Q38:AB38', Style(alignment=Alignment(horizontal='center'),
                                      border=Border(left=Side(border_style='thin', color=colors.BLACK),
                                                    right=Side(border_style='thin', color=colors.BLACK))), )
    style_range(ws, 'Q39:AB39', Style(alignment=Alignment(horizontal='center'),
                                      border=Border(left=Side(border_style='thin', color=colors.BLACK),
                                                    right=Side(border_style='thin', color=colors.BLACK))), )
    style_range(ws, 'Q40:AB40', Style(alignment=Alignment(horizontal='center'),
                                      border=Border(left=Side(border_style='thin', color=colors.BLACK),
                                                    right=Side(border_style='thin', color=colors.BLACK))), )
    style_range(ws, 'Q41:AB41', Style(alignment=Alignment(horizontal='center'),
                                      border=Border(left=Side(border_style='thin', color=colors.BLACK),
                                                    right=Side(border_style='thin', color=colors.BLACK))), )

    semcol = "B"
    subjnocol = "D"
    snamecol = "C"
    chacol = "E"
    chpcol = "H"
    gradecol = "K"
    qpacol = "N"
    cumqpa = 0
    cumcha = 0
    cumchp = 0
    cumqp = 0
    semgpa = 0
    semvar = 20
    semmmcount = 0
    # for cumulative gpa


    if z == 0:
        gpppp = 0
        gradeing = 0
        semcha = 0
        semchp = 0
        semqp = 0

        # semvar = 20
        a = semcol + str(semvar)
        ws[a] = (semname[0].get() + " " + semyear[0].get())
        subCcount = subcon[0].get()
        semcount = 1
        for q in range(0, int(subCcount)):
            ws[snamecol + str(semvar)] = subjname[t].get()
            if subjname[t].get() not in subject_list:
                ffw = open('subj_list.csv', 'a')
                ffw.write(",")
                ffw.write("\n")
                ffw.write(subjname[t].get())
                ffw.close()
            ws[subjnocol + str(semvar)] = subjcode[t].get()
            if subjcode[t].get() not in code_list:
                ffw1 = open('subject_code.csv', 'a')
                ffw1.write(",")
                ffw1.write("\n")
                ffw1.write(subjname[t].get())
                ffw1.close()
            ws[chacol + str(semvar)] = cha[t].get()
            semcha = semcha + int(cha[t].get())
            ws[chpcol + str(semvar)] = chp[t].get()
            semchp = semchp + int(chp[t].get())
            ws[gradecol + str(semvar)] = gr[t].get()
            ws[qpacol + str(semvar)] = gp[t].get()
            semqp = semqp + float(gp[t].get())
            gppp = float(gp[t].get())
            gradeing += gppp
            t += 1
            gpppp += 1
            semvar += 1
        gpa = (gradeing / (gpppp * 3))
        semgpa += gpa
        cumqpa += semgpa

        cumcha = cumcha + semcha
        cumchp = cumchp + semchp

        semvar += 1
        # for semester entry
        ws[semcol + str(semvar)] = semname[0].get() + " " + semyear[0].get()
        ws[snamecol + str(semvar)] = "Semester QPA"
        ws[subjnocol + str(semvar)] = gpa
        ws[chacol + str(semvar)] = semcha
        ws[chpcol + str(semvar)] = semchp
        ws[gradecol + str(semvar)] = ""
        ws[qpacol + str(semvar)] = semqp

        semvar += 1

        # for cumulative entry
        ws[semcol + str(semvar)] = semname[0].get() + " " + semyear[0].get()
        ws[snamecol + str(semvar)] = "Cumulative QPA"
        ws[subjnocol + str(semvar)] = cumqpa
        ws[chacol + str(semvar)] = cumcha
        ws[chpcol + str(semvar)] = cumchp
        ws[gradecol + str(semvar)] = ""
        ws[qpacol + str(semvar)] = cumqp

        semvar += 1

        # extra field
        ws['Q27'] = da.get()
        ws['Q21'] = da.get()
        ws['S28'] = datee.get()
        ws['S22'] = datee.get()







    else:
        for x in range(0, z + 1):
            gpppp = 0
            gradeing = 0
            semcha = 0
            semchp = 0
            semqp = 0
            semmmcount += 1
            semvar += 1

            subCcount = subcon[x].get()

            for q in range(0, int(subCcount)):
                a = semcol + str(semvar)
                ws[a] = semname[x].get() + " " + semyear[x].get()
                ws[snamecol + str(semvar)] = subjname[t].get()
                if subjname[t].get() not in subject_list:
                    ffw = open('subj_list.csv', 'a')
                    ffw.write(",")
                    ffw.write("\n")
                    ffw.write(subjname[t].get())
                    ffw.close()

                ws[subjnocol + str(semvar)] = subjcode[t].get()
                if subjcode[t].get() not in code_list:
                    ffw1 = open('subject_code.csv', 'a')
                    ffw1.write(",")
                    ffw1.write("\n")
                    ffw1.write(subjname[t].get())
                    ffw1.close()

                ws[chacol + str(semvar)] = cha[t].get()
                semcha = semcha + int(cha[t].get())
                ws[chpcol + str(semvar)] = chp[t].get()
                semchp = semchp + int(chp[t].get())
                ws[gradecol + str(semvar)] = gr[t].get()
                ws[qpacol + str(semvar)] = gp[t].get()
                semqp = semqp + float(gp[t].get())
                gppp = float(gp[t].get())
                gradeing += gppp
                t += 1
                gpppp += 1
                semvar += 1

            gpa = (gradeing / (gpppp * 3))
            semgpa = gpa
            cumqpa += semgpa
            cummqpa = cumqpa / semmmcount
            cumqp += semqp
            cumcha += semcha
            cumchp += semchp

            semvar += 1

            # for semester entry
            ws[semcol + str(semvar)] = semname[x].get() + " " + semyear[x].get()
            ws[snamecol + str(semvar)] = "Semester QPA"
            ws[subjnocol + str(semvar)] = gpa
            ws[chacol + str(semvar)] = semcha
            ws[chpcol + str(semvar)] = semchp
            ws[gradecol + str(semvar)] = ""
            ws[qpacol + str(semvar)] = semqp

            semvar += 1

            # for cumulative entry
            ws[semcol + str(semvar)] = semname[x].get() + " " + semyear[x].get()
            ws[snamecol + str(semvar)] = "Cumulative QPA"
            ws[subjnocol + str(semvar)] = cummqpa
            ws[chacol + str(semvar)] = cumcha
            ws[chpcol + str(semvar)] = cumchp
            ws[gradecol + str(semvar)] = ""
            ws[qpacol + str(semvar)] = cumqp

            semvar += 1




        # extra field
        ws['Q27'] = da.get()
        style_range(ws, 'Q27:AB27', Style(alignment=Alignment(horizontal='center'),
                                          border=Border(left=Side(border_style='thin', color=colors.BLACK),
                                                        right=Side(border_style='thin', color=colors.BLACK))), )

        ws['Q21'] = da.get()
        style_range(ws, 'Q21:AB21', Style(alignment=Alignment(horizontal='center'),
                                          border=Border(left=Side(border_style='thin', color=colors.BLACK),
                                                        right=Side(border_style='thin', color=colors.BLACK))), )
        ws['S28'] = datee.get()
        ws['S22'] = datee.get()

        wb.save(nameforfile + '.xlsx')
        complete1()
        style_range(ws, 'E77:N77', Style(alignment=Alignment(horizontal='center'),
                                         border=Border(bottom=Side(border_style='thin', color=colors.BLACK))), )


# Subject list
cdf = open('subj_list.csv')
csv_f1 = csv.reader(cdf)
subject_list = []
for row in csv_f1:
    subject_list.append(row[0])
subject_list = list(subject_list)
cdf.close()

# Course ID List
ccf = open('subject_code.csv')
csv_f = csv.reader(ccf)
code_list = []
for row in csv_f:
    code_list.append(row[0])
code_list = list(code_list)
ccf.close()


# Semester List
sem_list = ('FA', 'SP', 'SUMI', 'SUMII')


def complete1():
    global master
    fnnn = os.path.dirname(os.path.abspath(__file__))
    fnnn = "Transcript is generated " + fnnn
    tkMessageBox.showinfo("Transcript Generated", fnnn)
    result = tkMessageBox.askquestion("Generate new transcript", "Are You Sure?", icon='warning')
    if result == 'yes':
        master.destroy()

        master = Tk()

        # frame size
        master.geometry('1024x800')
        # window title
        master.title("TRANSCRIPTS")
        # student details
        showw()

        mainloop()
        # the main app
        master.mainloop()
    else:
        askagain()


def askagain():
    global master
    result = tkMessageBox.askquestion("Did you verify", "Are You Sure?", icon='warning')
    if result == 'yes':
        master.destroy()

        master = Tk()

        # frame size
        master.geometry('1024x800')
        # window title
        master.title("TRANSCRIPTS")
        # student details
        showw()

        mainloop()
        # the main app
        master.mainloop()
    else:
        master.destroy()


def showw():
    global i, b, c, sem_list, d, subcon, semcount, semname, semyear, subject, subjinc, subjcode, subjname, cha, chp, gr, gp, w, first, last, sd, ssn, variable, deg, major, matric, sex, acc, artnmus, bkg, cis, crj, eco, edu, eng, finnmgt, his, insnre, lan, law, mar, mgt, natsci, nur, psy, sec, spedra, socsci, tax, busele, laele, oc, hs, hsg, adg, sch, tc

    Label(master, text="First Name").grid(row=0, column=1)
    Label(master, text="Last Name").grid(row=0, column=2)
    Label(master, text="Date Of Birth").grid(row=0, column=3)
    Label(master, text="SSN").grid(row=0, column=4)
    Label(master, text="Campus").grid(row=0, column=5)

    first = Entry(master)
    first.grid(row=1, column=1)

    last = Entry(master)
    last.grid(row=1, column=2)
    sd = Entry(master)
    sd.grid(row=1, column=3)
    ssn = Entry(master)
    ssn.grid(row=1, column=4)
    variable = StringVar(master)
    variable.set("one")

    w = OptionMenu(master, variable, "NYC", "PLV", "BC")
    w.grid(row=1, column=5)

    Label(master, text="Degree").grid(row=0, column=6)
    Label(master, text="Major").grid(row=0, column=7)
    Label(master, text="Matric").grid(row=0, column=8)
    Label(master, text="Sex").grid(row=0, column=9)
    deg = Entry(master)
    deg.grid(row=1, column=6)
    major = Entry(master)
    major.grid(row=1, column=7)
    matric = Entry(master)
    matric.grid(row=1, column=8)
    sex = Entry(master)
    sex.grid(row=1, column=9)




    # grid variable
    i = 4

    Label(master, text="GRID").grid(row=i, column=1)

    i += 1
    Label(master, text="ACC").grid(row=i, column=0)
    Label(master, text="ART & MUS").grid(row=i, column=1)
    Label(master, text="BKG").grid(row=i, column=2)
    Label(master, text="CIS").grid(row=i, column=3)
    Label(master, text="CRJ").grid(row=i, column=4)
    Label(master, text="ECO").grid(row=i, column=5)
    Label(master, text="EDU").grid(row=i, column=6)
    Label(master, text="ENG").grid(row=i, column=7)
    Label(master, text="FIN & MGT").grid(row=i, column=8)
    Label(master, text="HIS").grid(row=i, column=9)
    Label(master, text="INS & RE").grid(row=i, column=10)
    Label(master, text="LAN").grid(row=i, column=11)

    i += 1
    acc = Entry(master)
    acc.grid(row=i, column=0)
    artnmus = Entry(master)
    artnmus.grid(row=i, column=1)
    bkg = Entry(master)
    bkg.grid(row=i, column=2)
    cis = Entry(master)
    cis.grid(row=i, column=3)
    crj = Entry(master)
    crj.grid(row=i, column=4)
    eco = Entry(master)
    eco.grid(row=i, column=5)
    edu = Entry(master)
    edu.grid(row=i, column=6)
    eng = Entry(master)
    eng.grid(row=i, column=7)
    finnmgt = Entry(master)
    finnmgt.grid(row=i, column=8)
    his = Entry(master)
    his.grid(row=i, column=9)
    insnre = Entry(master)
    insnre.grid(row=i, column=10)
    lan = Entry(master)
    lan.grid(row=i, column=11)
    law = Entry(master)

    i += 1

    Label(master, text="LAW").grid(row=i, column=0)
    Label(master, text="MAR").grid(row=i, column=1)
    Label(master, text="MAT").grid(row=i, column=2)
    Label(master, text="NAT SCI").grid(row=i, column=3)
    Label(master, text="NUR").grid(row=i, column=4)
    Label(master, text="PSY").grid(row=i, column=5)
    Label(master, text="SEC").grid(row=i, column=6)
    Label(master, text="SPE DRA").grid(row=i, column=7)
    Label(master, text="SOC SCI").grid(row=i, column=8)
    Label(master, text="TAX").grid(row=i, column=9)
    Label(master, text="BUS ELE").grid(row=i, column=10)
    Label(master, text="LA ELE").grid(row=i, column=11)

    i += 1

    law.grid(row=i, column=0)
    mar = Entry(master)
    mar.grid(row=i, column=1)
    mgt = Entry(master)
    mgt.grid(row=i, column=2)
    natsci = Entry(master)
    natsci.grid(row=i, column=3)
    nur = Entry(master)
    nur.grid(row=i, column=4)
    psy = Entry(master)
    psy.grid(row=i, column=5)
    sec = Entry(master)
    sec.grid(row=i, column=6)
    spedra = Entry(master)
    spedra.grid(row=i, column=7)
    socsci = Entry(master)
    socsci.grid(row=i, column=8)
    tax = Entry(master)
    tax.grid(row=i, column=9)
    busele = Entry(master)
    busele.grid(row=i, column=10)
    laele = Entry(master)
    laele.grid(row=i, column=11)

    # i+=1
    # Label(master, text="High School & Other College").grid(row=i, column=0)
    i += 1
    Label(master, text="Other College").grid(row=i, column=0)
    Label(master, text="High School").grid(row=i, column=1)
    Label(master, text="High School Graduation").grid(row=i, column=2)
    Label(master, text="Advanced Degree Granted").grid(row=i, column=3)
    i += 1
    oc = Entry(master)
    oc.grid(row=i, column=0)
    hs = Entry(master)
    hs.grid(row=i, column=1)
    hsg = Entry(master)
    hsg.grid(row=i, column=2)
    adg = Entry(master)
    adg.grid(row=i, column=3)

    i += 1

    Label(master, text="School").grid(row=i, column=0)
    Label(master, text="Total Credits").grid(row=i, column=1)
    i += 1
    sch = Entry(master)
    sch.grid(row=i, column=0)
    tc = Entry(master)
    tc.grid(row=i, column=1)

    b = Button(master, text="Add Semester", command=callsem)
    b.grid(row=(i + 1), column=2)
    c = Button(master, text="Add Subject", command=callback)
    c.grid(row=(i + 1), column=3)
    d = Button(master, text="Enter Final Detail", command=calldegree)
    d.grid(row=(i + 1), column=4)


master = Tk()
#
# # frame size
master.geometry('1024x800')
# # window title
master.title("TRANSCRIPTS")
# student details
showw()

master.mainloop()
